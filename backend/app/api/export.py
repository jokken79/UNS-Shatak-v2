"""
Export API - Exportación de datos a CSV y Excel
"""

from typing import List, Optional
from uuid import UUID
from datetime import datetime, date
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from sqlalchemy import and_, func
import csv
import io
from fastapi.responses import StreamingResponse

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from app.core.database import get_db
from app.core.security import get_current_user
from app.models.models import Apartment, Employee, ApartmentAssignment, Factory, User

router = APIRouter(prefix="/export", tags=["Export (エクスポート)"])


def format_value(value):
    """Format values for export"""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Sí" if value else "No"
    return str(value)


@router.get("/occupancy/csv")
async def export_occupancy_csv(
    apartment_id: Optional[UUID] = None,
    month: Optional[int] = None,
    year: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Export occupancy data to CSV
    Includes: Apartment code, employee code, name, move-in date, color, status
    """
    query = db.query(
        Apartment.apartment_code,
        Apartment.name.label("apartment_name"),
        Apartment.capacity,
        Employee.employee_code,
        Employee.full_name_roman,
        Employee.full_name_kanji,
        Employee.email,
        Employee.phone,
        Factory.name.label("factory_name"),
        ApartmentAssignment.move_in_date,
        ApartmentAssignment.move_out_date,
        ApartmentAssignment.is_current,
        ApartmentAssignment.is_recent,
        ApartmentAssignment.assigned_color,
        ApartmentAssignment.monthly_charge
    ).select_from(Apartment).outerjoin(
        ApartmentAssignment, Apartment.id == ApartmentAssignment.apartment_id
    ).outerjoin(
        Employee, ApartmentAssignment.employee_id == Employee.id
    ).outerjoin(
        Factory, Employee.factory_id == Factory.id
    ).filter(Apartment.is_active == True)

    if apartment_id:
        query = query.filter(Apartment.id == apartment_id)

    if month and year:
        query = query.filter(
            func.extract('month', ApartmentAssignment.move_in_date) == month,
            func.extract('year', ApartmentAssignment.move_in_date) == year
        )

    results = query.order_by(Apartment.apartment_code, Employee.employee_code).all()

    # Create CSV
    output = io.StringIO()
    writer = csv.writer(output, delimiter=',', quoting=csv.QUOTE_ALL)

    # Headers
    headers = [
        "Código Apartamento",
        "Nombre Apartamento",
        "Capacidad",
        "Código Empleado",
        "Nombre (Romaji)",
        "Nombre (Kanji)",
        "Email",
        "Teléfono",
        "Fábrica",
        "Fecha Ingreso",
        "Fecha Salida",
        "Actual",
        "Nuevo",
        "Color",
        "Renta Mensual"
    ]
    writer.writerow(headers)

    # Data
    for row in results:
        writer.writerow([
            format_value(row.apartment_code),
            format_value(row.apartment_name),
            format_value(row.capacity),
            format_value(row.employee_code),
            format_value(row.full_name_roman),
            format_value(row.full_name_kanji),
            format_value(row.email),
            format_value(row.phone),
            format_value(row.factory_name),
            format_value(row.move_in_date),
            format_value(row.move_out_date),
            format_value(row.is_current),
            format_value(row.is_recent),
            format_value(row.assigned_color),
            format_value(row.monthly_charge)
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=occupancy_export.csv"}
    )


@router.get("/occupancy/excel")
async def export_occupancy_excel(
    apartment_id: Optional[UUID] = None,
    month: Optional[int] = None,
    year: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Export occupancy data to Excel with colors and formatting
    Includes: Apartment code, employee code, name, move-in date, color, status
    """
    if not EXCEL_AVAILABLE:
        raise HTTPException(
            status_code=400,
            detail="Excel export requires openpyxl. Install: pip install openpyxl"
        )

    query = db.query(
        Apartment.apartment_code,
        Apartment.name.label("apartment_name"),
        Apartment.capacity,
        Employee.employee_code,
        Employee.full_name_roman,
        Employee.full_name_kanji,
        Employee.email,
        Employee.phone,
        Factory.name.label("factory_name"),
        ApartmentAssignment.move_in_date,
        ApartmentAssignment.move_out_date,
        ApartmentAssignment.is_current,
        ApartmentAssignment.is_recent,
        ApartmentAssignment.assigned_color,
        ApartmentAssignment.monthly_charge
    ).select_from(Apartment).outerjoin(
        ApartmentAssignment, Apartment.id == ApartmentAssignment.apartment_id
    ).outerjoin(
        Employee, ApartmentAssignment.employee_id == Employee.id
    ).outerjoin(
        Factory, Employee.factory_id == Factory.id
    ).filter(Apartment.is_active == True)

    if apartment_id:
        query = query.filter(Apartment.id == apartment_id)

    if month and year:
        query = query.filter(
            func.extract('month', ApartmentAssignment.move_in_date) == month,
            func.extract('year', ApartmentAssignment.move_in_date) == year
        )

    results = query.order_by(Apartment.apartment_code, Employee.employee_code).all()

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ocupación"

    # Headers
    headers = [
        "Código Apartamento",
        "Nombre Apartamento",
        "Capacidad",
        "Código Empleado",
        "Nombre (Romaji)",
        "Nombre (Kanji)",
        "Email",
        "Teléfono",
        "Fábrica",
        "Fecha Ingreso",
        "Fecha Salida",
        "Actual",
        "Nuevo",
        "Color",
        "Renta Mensual"
    ]

    # Write headers with formatting
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Set column widths
    col_widths = [18, 25, 10, 15, 20, 20, 25, 15, 20, 15, 15, 8, 8, 12, 15]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Write data
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row_idx, row_data in enumerate(results, 2):
        # Determine row color based on is_recent
        if row_data.is_recent:
            row_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
            row_font = Font(bold=True)
        else:
            row_fill = None
            row_font = None

        for col_idx, value in enumerate([
            row_data.apartment_code,
            row_data.apartment_name,
            row_data.capacity,
            row_data.employee_code,
            row_data.full_name_roman,
            row_data.full_name_kanji,
            row_data.email,
            row_data.phone,
            row_data.factory_name,
            row_data.move_in_date,
            row_data.move_out_date,
            row_data.is_current,
            row_data.is_recent,
            row_data.assigned_color,
            row_data.monthly_charge
        ], 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=format_value(value))
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            if row_fill:
                cell.fill = row_fill
            if row_font:
                cell.font = row_font

            # Color cell for assigned_color column
            if col_idx == 14 and row_data.assigned_color:
                try:
                    # Remove # from hex color
                    hex_color = row_data.assigned_color.lstrip('#')
                    cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                    # Change font color to white for visibility
                    cell.font = Font(color="FFFFFF", bold=True)
                except:
                    pass

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=occupancy_export.xlsx"}
    )


@router.get("/occupancy/summary")
async def export_occupancy_summary(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get summary of occupancy for export
    Returns JSON with all apartments and their residents
    """
    apartments = db.query(Apartment).filter(Apartment.is_active == True).all()

    summary = []
    for apartment in apartments:
        assignments = db.query(
            ApartmentAssignment,
            Employee,
            Factory
        ).select_from(ApartmentAssignment).join(
            Employee
        ).outerjoin(Factory).filter(
            ApartmentAssignment.apartment_id == apartment.id,
            ApartmentAssignment.is_current == True
        ).all()

        residents = []
        for assignment, employee, factory in assignments:
            residents.append({
                "employee_code": employee.employee_code,
                "employee_id": str(employee.id),
                "full_name_roman": employee.full_name_roman,
                "full_name_kanji": employee.full_name_kanji,
                "email": employee.email,
                "phone": employee.phone,
                "factory_name": factory.name if factory else None,
                "move_in_date": assignment.move_in_date.isoformat() if assignment.move_in_date else None,
                "is_recent": assignment.is_recent,
                "assigned_color": assignment.assigned_color,
                "monthly_charge": float(assignment.monthly_charge) if assignment.monthly_charge else None
            })

        summary.append({
            "apartment_code": apartment.apartment_code,
            "apartment_id": str(apartment.id),
            "apartment_name": apartment.name,
            "address": apartment.address,
            "capacity": apartment.capacity,
            "current_occupants": apartment.current_occupants,
            "status": apartment.status,
            "monthly_rent": float(apartment.monthly_rent) if apartment.monthly_rent else None,
            "pricing_type": apartment.pricing_type,
            "residents": residents,
            "occupancy_rate": round((len(residents) / apartment.capacity * 100) if apartment.capacity > 0 else 0, 2)
        })

    return {
        "export_date": datetime.now().isoformat(),
        "total_apartments": len(summary),
        "total_residents": sum(len(apt["residents"]) for apt in summary),
        "apartments": summary
    }
